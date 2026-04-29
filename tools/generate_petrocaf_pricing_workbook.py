"""
PETROCAF Pricing Workbook Generator
===================================

This script generates a complete Excel pricing workbook for PETROCAF.
It is designed as a deterministic, auditable BOQ pricing starter engine.

Output file:
    PETROCAF_Pricing_Master_Workbook.xlsx

Run:
    pip install openpyxl
    python tools/generate_petrocaf_pricing_workbook.py

Important rules:
- No fabricated source-backed rates.
- Every placeholder cost/rate is tagged as [ESTIMATED] or [MISSING].
- AI may support classification/review only; final pricing logic must remain rule-based and auditable.
- The workbook is a pricing engine template, not a final commercial quote.
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

OUTPUT_FILE = "PETROCAF_Pricing_Master_Workbook.xlsx"

COMPANY = "PETROCAF - Petroleum Construction All Facilities LLC"
CURRENCY = "EGP"
VERSION = "V1.0"

# -----------------------------
# Styling helpers
# -----------------------------

FILL_NAVY = "1F4E79"
FILL_DARK = "243746"
FILL_GOLD = "C9A227"
FILL_LIGHT = "EAF2F8"
FILL_WARN = "FFF2CC"
FILL_BAD = "F4CCCC"
FILL_GOOD = "D9EAD3"
WHITE = "FFFFFF"
BLACK = "000000"
GREY = "D9E1F2"

THIN_GREY = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)


def style_title(ws: Worksheet, title: str, subtitle: str | None = None) -> None:
    ws.merge_cells("A1:J1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=18, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=FILL_NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws.merge_cells("A2:J2")
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, size=11, color=BLACK)
        ws["A2"].fill = PatternFill("solid", fgColor=FILL_LIGHT)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22


def write_table(ws: Worksheet, start_row: int, headers: Sequence[str], rows: Iterable[Sequence[object]], table_name: str) -> int:
    for col, header in enumerate(headers, 1):
        c = ws.cell(start_row, col, header)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=FILL_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    end_row = start_row
    for r_idx, row in enumerate(rows, start_row + 1):
        end_row = r_idx
        for c_idx, value in enumerate(row, 1):
            c = ws.cell(r_idx, c_idx, value)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER

    end_col = len(headers)
    ref = f"A{start_row}:{get_column_letter(end_col)}{max(end_row, start_row + 1)}"
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    return end_row


def autofit(ws: Worksheet, min_width: int = 10, max_width: int = 45) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_width
        for cell in col:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def freeze_and_filter(ws: Worksheet, row: int = 4) -> None:
    ws.freeze_panes = f"A{row}"


def hide_grid(ws: Worksheet) -> None:
    ws.sheet_view.showGridLines = False


def add_validation(ws: Worksheet, cell_range: str, values: Sequence[str]) -> None:
    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


# -----------------------------
# Workbook sheet builders
# -----------------------------


def build_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README_BASIS"
    hide_grid(ws)
    style_title(ws, "PETROCAF Pricing Engine - Basis", "Deterministic, auditable and execution-oriented pricing workbook")

    headers = ["Section", "Content"]
    rows = [
        ["Company", COMPANY],
        ["Workbook Version", VERSION],
        ["Currency", CURRENCY],
        ["Primary Use", "BOQ pricing, cost build-up, review flags, tender assumptions and executive summary."],
        ["Pricing Principle", "Rule-based calculation. No invented source-backed data. Use [MISSING], [ESTIMATED] and [SOURCE-BACKED] tags."],
        ["AI Boundary", "AI may classify BOQ descriptions and highlight gaps only. AI must not invent rates or productivity figures."],
        ["Estimator Action", "Replace all [MISSING] and [ESTIMATED] values with verified supplier quotes, approved company norms or project-specific assumptions."],
        ["Governance", "Every rate, productivity and factor must have owner, date, source status and revision trace."],
        ["Risk Note", "Do not submit commercial offer before reviewing scope exclusions, site conditions, client specs, taxes, logistics and payment terms."],
    ]
    write_table(ws, 4, headers, rows, "tbl_readme_basis")
    autofit(ws)


def build_project_master(wb: Workbook) -> None:
    ws = wb.create_sheet("PROJECT_MASTER")
    hide_grid(ws)
    style_title(ws, "Project Master", "Control project metadata and pricing basis")

    headers = ["Field", "Value", "Status", "Notes"]
    rows = [
        ["Project Name", "[MISSING]", "MISSING", "Enter tender/project name."],
        ["Client", "[MISSING]", "MISSING", "Enter client/company."],
        ["Location", "Egypt", "ESTIMATED", "Replace with exact site."],
        ["Tender No.", "[MISSING]", "MISSING", "Client tender reference."],
        ["Pricing Date", "=TODAY()", "AUTO", "Workbook date."],
        ["Currency", CURRENCY, "DEFAULT", "Change if required."],
        ["Base Working Hours/Day", 10, "ESTIMATED", "Adjust by project calendar."],
        ["Working Days/Month", 26, "ESTIMATED", "Adjust by actual roster."],
        ["Default OH %", 0.12, "ESTIMATED", "Company overhead."],
        ["Default Profit %", 0.10, "ESTIMATED", "Management decision."],
        ["Default Contingency %", 0.07, "ESTIMATED", "Risk-based; revise by scope."],
        ["VAT / Tax %", 0.14, "ESTIMATED", "Confirm tax treatment."],
        ["Prepared By", "PETROCAF Estimation Team", "DEFAULT", "Update owner."],
        ["Reviewed By", "[MISSING]", "MISSING", "Technical/commercial reviewer."],
        ["Approved By", "[MISSING]", "MISSING", "Management approval."],
    ]
    write_table(ws, 4, headers, rows, "tbl_project_master")
    autofit(ws)
    ws.column_dimensions["B"].width = 28


def build_boq_input(wb: Workbook) -> None:
    ws = wb.create_sheet("BOQ_INPUT")
    hide_grid(ws)
    style_title(ws, "BOQ Input", "Paste client BOQ items here then map to activity codes")

    headers = [
        "Item_ID", "WBS", "Discipline", "Scope", "Activity_Code", "Description", "UOM", "Qty",
        "Dia_Inch", "Sch_Class", "Material", "Service", "Location", "Execution_Condition",
        "Mapped_By", "Source_Status", "Estimator_Notes"
    ]
    rows = [
        ["BOQ-001", "PIP-001", "Mechanical", "Piping", "PIP_SPOOL_FAB_CS", "Fabrication of carbon steel pipe spool including fit-up and welding", "INCH-DIA", 120, 6, "STD", "CS", "Process", "Workshop", "Normal", "Manual", "ESTIMATED", "Replace with actual BOQ."],
        ["BOQ-002", "PIP-002", "Mechanical", "Piping", "PIP_INSTALL_ABOVEGROUND", "Installation of aboveground CS piping including supports alignment", "INCH-DIA", 180, 4, "STD", "CS", "Utility", "Site", "Congested", "Manual", "ESTIMATED", "Check scaffolding and access."],
        ["BOQ-003", "HYD-001", "Mechanical", "Hydrotest", "HYD_TEST_PACKAGE", "Hydrotest piping test package including filling pressurizing and reinstatement", "PACKAGE", 3, "", "", "CS", "Water", "Site", "Normal", "Manual", "ESTIMATED", "Add water disposal and test blinds."],
        ["BOQ-004", "CIV-001", "Civil", "Concrete", "CIV_CONCRETE_FOUNDATION", "RC concrete foundation including pouring and finishing", "M3", 25, "", "", "Concrete", "", "Site", "Normal", "Manual", "ESTIMATED", "Civil rates are placeholders."],
        ["BOQ-005", "EI-001", "E&I", "Cable", "EI_CABLE_PULLING", "Cable pulling in tray/conduit including tagging", "M", 350, "", "", "Cu Cable", "Electrical", "Site", "Normal", "Manual", "ESTIMATED", "Confirm cable size."],
    ]
    write_table(ws, 4, headers, rows, "tbl_boq_input")
    add_validation(ws, "C5:C500", ["Mechanical", "Civil", "E&I", "Painting", "Insulation", "Scaffolding", "Other"])
    add_validation(ws, "G5:G500", ["EA", "M", "M2", "M3", "KG", "TON", "INCH-DIA", "JOINT", "PACKAGE", "LS"])
    add_validation(ws, "P5:P500", ["MISSING", "ESTIMATED", "SOURCE-BACKED", "CLIENT-BOQ", "SUPPLIER-QUOTE"])
    freeze_and_filter(ws, 5)
    autofit(ws)


def build_activity_library(wb: Workbook) -> None:
    ws = wb.create_sheet("ACTIVITY_LIBRARY")
    hide_grid(ws)
    style_title(ws, "Activity Library", "Standard activity dictionary used by BOQ mapping")

    headers = [
        "Activity_Code", "Discipline", "Scope", "Activity_Name", "Default_UOM", "Pricing_Basis",
        "Default_Productivity_Code", "Default_Crew_Code", "Default_Equipment_Code", "Default_Material_Code",
        "Risk_Class", "Source_Status", "Notes"
    ]
    rows = [
        ["PIP_SPOOL_FAB_CS", "Mechanical", "Piping", "CS pipe spool fabrication", "INCH-DIA", "Labor + equipment + consumables", "PROD_PIP_FAB_ID", "CREW_PIP_FAB", "EQ_FAB_SHOP", "MAT_WELD_CONS", "Medium", "ESTIMATED", "Workshop fabrication basis."],
        ["PIP_INSTALL_ABOVEGROUND", "Mechanical", "Piping", "Aboveground piping installation", "INCH-DIA", "Labor + lifting + tools", "PROD_PIP_ERECT_ID", "CREW_PIP_ERECT", "EQ_LIFTING_LIGHT", "MAT_MISC_INSTALL", "High", "ESTIMATED", "Depends on access/congestion."],
        ["PIP_SUPPORT_INSTALL", "Mechanical", "Piping", "Pipe support installation", "KG", "Labor + equipment", "PROD_SUPPORT_KG", "CREW_STEEL_INSTALL", "EQ_LIFTING_LIGHT", "MAT_MISC_INSTALL", "Medium", "ESTIMATED", "Confirm galvanized/painted support basis."],
        ["HYD_TEST_PACKAGE", "Mechanical", "Hydrotest", "Hydrotest package", "PACKAGE", "Labor + pump + water + consumables", "PROD_HYD_PACKAGE", "CREW_HYD_TEST", "EQ_HYD_PUMP", "MAT_HYD_CONS", "High", "ESTIMATED", "Include test packs, calibration, reinstatement."],
        ["TANK_REPAIR_PLATE", "Mechanical", "Tank", "Tank plate repair", "M2", "Labor + welding + lifting", "PROD_TANK_PLATE_M2", "CREW_TANK_REPAIR", "EQ_FAB_SITE", "MAT_STEEL_PLATE", "High", "ESTIMATED", "API 653 review required if applicable."],
        ["CIV_EARTHWORK", "Civil", "Earthwork", "Excavation/backfilling", "M3", "Labor + equipment", "PROD_CIV_EARTH_M3", "CREW_CIVIL", "EQ_EXCAVATION", "MAT_NONE", "Medium", "ESTIMATED", "Soil and disposal route required."],
        ["CIV_CONCRETE_FOUNDATION", "Civil", "Concrete", "RC concrete foundation", "M3", "Labor + equipment + material", "PROD_CONC_M3", "CREW_CIVIL", "EQ_CONCRETE", "MAT_CONCRETE", "Medium", "ESTIMATED", "Include formwork/rebar separately if required."],
        ["EI_CABLE_PULLING", "E&I", "Cable", "Cable pulling", "M", "Labor + tools", "PROD_CABLE_M", "CREW_EI_CABLE", "EQ_EI_TOOLS", "MAT_MISC_EI", "Medium", "ESTIMATED", "Productivity depends on tray route and cable size."],
        ["EI_INSTR_CALIBRATION", "E&I", "Instrument", "Instrument calibration", "EA", "Labor + tools", "PROD_INST_CAL_EA", "CREW_INST_CAL", "EQ_CALIBRATION", "MAT_MISC_EI", "Medium", "ESTIMATED", "Certified calibration equipment needed."],
        ["PAINT_BLAST_PAINT", "Painting", "Coating", "Blasting and painting", "M2", "Labor + equipment + consumables", "PROD_PAINT_M2", "CREW_PAINT", "EQ_PAINT", "MAT_PAINT_SYSTEM", "High", "MISSING", "Requires paint system, DFT, surface prep grade."],
    ]
    write_table(ws, 4, headers, rows, "tbl_activity_library")
    autofit(ws)


def build_productivity_library(wb: Workbook) -> None:
    ws = wb.create_sheet("PRODUCTIVITY_LIBRARY")
    hide_grid(ws)
    style_title(ws, "Productivity Library", "Editable productivity norms with source status")

    headers = [
        "Productivity_Code", "Activity_Code", "UOM", "Base_Productivity_Per_Day", "Crew_Code",
        "Working_Hours_Day", "Condition_Factor", "Location_Factor", "Complexity_Factor", "Adjusted_Productivity_Per_Day",
        "Source_Status", "Reference", "Notes"
    ]
    rows = [
        ["PROD_PIP_FAB_ID", "PIP_SPOOL_FAB_CS", "INCH-DIA", 30, "CREW_PIP_FAB", 10, 1.00, 1.00, 1.00, "=D5*G5*H5*I5", "ESTIMATED", "Company norm required", "Replace after calibration."],
        ["PROD_PIP_ERECT_ID", "PIP_INSTALL_ABOVEGROUND", "INCH-DIA", 24, "CREW_PIP_ERECT", 10, 0.80, 0.85, 0.90, "=D6*G6*H6*I6", "ESTIMATED", "Company norm required", "Congested site factor applied."],
        ["PROD_SUPPORT_KG", "PIP_SUPPORT_INSTALL", "KG", 450, "CREW_STEEL_INSTALL", 10, 0.90, 0.90, 0.95, "=D7*G7*H7*I7", "ESTIMATED", "Company norm required", "Depends on support complexity."],
        ["PROD_HYD_PACKAGE", "HYD_TEST_PACKAGE", "PACKAGE", 1, "CREW_HYD_TEST", 10, 0.80, 0.90, 0.85, "=D8*G8*H8*I8", "ESTIMATED", "Company norm required", "One package/day may be aggressive."],
        ["PROD_TANK_PLATE_M2", "TANK_REPAIR_PLATE", "M2", 8, "CREW_TANK_REPAIR", 10, 0.85, 0.85, 0.80, "=D9*G9*H9*I9", "MISSING", "To be sourced", "Requires tank repair method."],
        ["PROD_CIV_EARTH_M3", "CIV_EARTHWORK", "M3", 80, "CREW_CIVIL", 10, 1.00, 0.95, 0.90, "=D10*G10*H10*I10", "ESTIMATED", "Company norm required", "Depends on equipment and disposal."],
        ["PROD_CONC_M3", "CIV_CONCRETE_FOUNDATION", "M3", 18, "CREW_CIVIL", 10, 1.00, 1.00, 0.95, "=D11*G11*H11*I11", "ESTIMATED", "Company norm required", "Check formwork/rebar split."],
        ["PROD_CABLE_M", "EI_CABLE_PULLING", "M", 180, "CREW_EI_CABLE", 10, 0.90, 0.90, 0.90, "=D12*G12*H12*I12", "ESTIMATED", "Company norm required", "Cable size affects productivity."],
        ["PROD_INST_CAL_EA", "EI_INSTR_CALIBRATION", "EA", 12, "CREW_INST_CAL", 10, 1.00, 1.00, 0.95, "=D13*G13*H13*I13", "ESTIMATED", "Company norm required", "Certified calibrator needed."],
        ["PROD_PAINT_M2", "PAINT_BLAST_PAINT", "M2", 55, "CREW_PAINT", 10, 0.90, 0.90, 0.85, "=D14*G14*H14*I14", "MISSING", "Paint subcontractor quote", "Need paint system."],
    ]
    write_table(ws, 4, headers, rows, "tbl_productivity_library")
    autofit(ws)


def build_crew_library(wb: Workbook) -> None:
    ws = wb.create_sheet("CREW_LIBRARY")
    hide_grid(ws)
    style_title(ws, "Crew Library", "Crew templates and labor rates")

    headers = [
        "Crew_Code", "Role", "Qty", "Daily_Rate", "Rate_Status", "Hours_Day", "Cost_Per_Day", "Manhours_Per_Day", "Notes"
    ]
    rows = [
        ["CREW_PIP_FAB", "Piping Supervisor", 1, 900, "ESTIMATED", 10, "=C5*D5", "=C5*F5", ""],
        ["CREW_PIP_FAB", "Pipe Fitter", 3, 650, "ESTIMATED", 10, "=C6*D6", "=C6*F6", ""],
        ["CREW_PIP_FAB", "Welder", 2, 750, "ESTIMATED", 10, "=C7*D7", "=C7*F7", ""],
        ["CREW_PIP_FAB", "Helper", 3, 400, "ESTIMATED", 10, "=C8*D8", "=C8*F8", ""],
        ["CREW_PIP_ERECT", "Mechanical Supervisor", 1, 900, "ESTIMATED", 10, "=C9*D9", "=C9*F9", ""],
        ["CREW_PIP_ERECT", "Pipe Fitter", 4, 650, "ESTIMATED", 10, "=C10*D10", "=C10*F10", ""],
        ["CREW_PIP_ERECT", "Rigger", 2, 600, "ESTIMATED", 10, "=C11*D11", "=C11*F11", ""],
        ["CREW_PIP_ERECT", "Helper", 4, 400, "ESTIMATED", 10, "=C12*D12", "=C12*F12", ""],
        ["CREW_HYD_TEST", "Hydrotest Engineer", 1, 1200, "ESTIMATED", 10, "=C13*D13", "=C13*F13", ""],
        ["CREW_HYD_TEST", "Hydrotest Technician", 3, 650, "ESTIMATED", 10, "=C14*D14", "=C14*F14", ""],
        ["CREW_HYD_TEST", "Helper", 3, 400, "ESTIMATED", 10, "=C15*D15", "=C15*F15", ""],
        ["CREW_STEEL_INSTALL", "Steel Supervisor", 1, 900, "ESTIMATED", 10, "=C16*D16", "=C16*F16", ""],
        ["CREW_STEEL_INSTALL", "Fabricator", 3, 650, "ESTIMATED", 10, "=C17*D17", "=C17*F17", ""],
        ["CREW_STEEL_INSTALL", "Welder", 1, 750, "ESTIMATED", 10, "=C18*D18", "=C18*F18", ""],
        ["CREW_CIVIL", "Civil Supervisor", 1, 850, "ESTIMATED", 10, "=C19*D19", "=C19*F19", ""],
        ["CREW_CIVIL", "Mason/Carpenter", 4, 550, "ESTIMATED", 10, "=C20*D20", "=C20*F20", ""],
        ["CREW_CIVIL", "Helper", 6, 400, "ESTIMATED", 10, "=C21*D21", "=C21*F21", ""],
        ["CREW_EI_CABLE", "E&I Supervisor", 1, 900, "ESTIMATED", 10, "=C22*D22", "=C22*F22", ""],
        ["CREW_EI_CABLE", "Electrician", 3, 650, "ESTIMATED", 10, "=C23*D23", "=C23*F23", ""],
        ["CREW_EI_CABLE", "Helper", 4, 400, "ESTIMATED", 10, "=C24*D24", "=C24*F24", ""],
        ["CREW_INST_CAL", "Instrument Technician", 2, 750, "ESTIMATED", 10, "=C25*D25", "=C25*F25", ""],
        ["CREW_INST_CAL", "Helper", 1, 400, "ESTIMATED", 10, "=C26*D26", "=C26*F26", ""],
        ["CREW_PAINT", "Painting Supervisor", 1, 850, "MISSING", 10, "=C27*D27", "=C27*F27", "Subcontract rate recommended."],
        ["CREW_PAINT", "Painter/Blaster", 4, 600, "MISSING", 10, "=C28*D28", "=C28*F28", "Subcontract rate recommended."],
    ]
    write_table(ws, 4, headers, rows, "tbl_crew_library")
    autofit(ws)


def build_cost_libraries(wb: Workbook) -> None:
    mat = wb.create_sheet("MATERIAL_LIBRARY")
    hide_grid(mat)
    style_title(mat, "Material Library", "Material, consumables and subcontract placeholders")
    mat_headers = ["Material_Code", "Description", "UOM", "Unit_Cost", "Waste_%", "Source_Status", "Reference", "Notes"]
    mat_rows = [
        ["MAT_NONE", "No material included", "EA", 0, 0, "DEFAULT", "N/A", "For labor-only items."],
        ["MAT_WELD_CONS", "Welding consumables per inch-dia", "INCH-DIA", 25, 0.05, "ESTIMATED", "Supplier quote required", "Electrodes/wire/gas/grinding discs."],
        ["MAT_MISC_INSTALL", "Miscellaneous installation consumables", "INCH-DIA", 15, 0.05, "ESTIMATED", "Supplier quote required", "Bolts, gaskets, small tools excluded unless stated."],
        ["MAT_HYD_CONS", "Hydrotest consumables per package", "PACKAGE", 4500, 0.03, "ESTIMATED", "Supplier quote required", "Water, temporary gaskets, blinds, fittings."],
        ["MAT_STEEL_PLATE", "Steel plate supply", "M2", 0, 0.08, "MISSING", "Supplier quote required", "Do not price without thickness/spec."],
        ["MAT_CONCRETE", "Ready mix concrete", "M3", 0, 0.03, "MISSING", "Supplier quote required", "Grade and delivery location required."],
        ["MAT_MISC_EI", "E&I consumables", "M", 8, 0.03, "ESTIMATED", "Supplier quote required", "Cable ties, tags, lugs excluded unless confirmed."],
        ["MAT_PAINT_SYSTEM", "Paint/coating system", "M2", 0, 0.08, "MISSING", "Paint vendor quote", "System, DFT and surface prep required."],
    ]
    write_table(mat, 4, mat_headers, mat_rows, "tbl_material_library")
    autofit(mat)

    eq = wb.create_sheet("EQUIPMENT_LIBRARY")
    hide_grid(eq)
    style_title(eq, "Equipment Library", "Equipment spreads and daily cost assumptions")
    eq_headers = ["Equipment_Code", "Description", "UOM", "Daily_Cost", "Utilization_Factor", "Source_Status", "Reference", "Notes"]
    eq_rows = [
        ["EQ_NONE", "No equipment", "DAY", 0, 0, "DEFAULT", "N/A", ""],
        ["EQ_FAB_SHOP", "Workshop tools and welding machines", "DAY", 1800, 0.75, "ESTIMATED", "Company/supplier required", "Allocated spread."],
        ["EQ_LIFTING_LIGHT", "Light lifting / chain blocks / forklift allowance", "DAY", 2500, 0.65, "ESTIMATED", "Supplier required", "Crane excluded unless required."],
        ["EQ_HYD_PUMP", "Hydrotest pump, manifold and hoses", "DAY", 3500, 0.80, "ESTIMATED", "Supplier required", "Calibration cert needed."],
        ["EQ_FAB_SITE", "Site fabrication equipment", "DAY", 2500, 0.70, "ESTIMATED", "Company/supplier required", ""],
        ["EQ_EXCAVATION", "Excavator / loader spread", "DAY", 6000, 0.80, "ESTIMATED", "Supplier required", "Disposal trucks excluded unless added."],
        ["EQ_CONCRETE", "Concrete tools / vibrator / small equipment", "DAY", 1800, 0.70, "ESTIMATED", "Supplier required", "Pump excluded unless needed."],
        ["EQ_EI_TOOLS", "Electrical hand tools and testing tools", "DAY", 700, 0.70, "ESTIMATED", "Company norm required", ""],
        ["EQ_CALIBRATION", "Calibration equipment", "DAY", 1200, 0.65, "ESTIMATED", "Calibration vendor required", "Valid certificates required."],
        ["EQ_PAINT", "Blasting/painting equipment", "DAY", 0, 0.75, "MISSING", "Subcontractor quote", "Compressor/pot/hoses/PPE."],
    ]
    write_table(eq, 4, eq_headers, eq_rows, "tbl_equipment_library")
    autofit(eq)


def build_indirects_and_calibration(wb: Workbook) -> None:
    ws = wb.create_sheet("INDIRECTS_RISK")
    hide_grid(ws)
    style_title(ws, "Indirects and Risk", "Project-level adders and pricing governance")
    headers = ["Cost_Item", "Basis", "Value", "Source_Status", "Notes"]
    rows = [
        ["Site Management", "% of direct cost", 0.06, "ESTIMATED", "Construction manager, engineers, admin."],
        ["HSE / QAQC", "% of direct cost", 0.025, "ESTIMATED", "Adjust for client requirements."],
        ["Temporary Facilities", "% of direct cost", 0.03, "ESTIMATED", "Offices, stores, utilities."],
        ["Mobilization/Demobilization", "% of direct cost", 0.04, "ESTIMATED", "Depends on location."],
        ["Consumables Not Itemized", "% of direct cost", 0.02, "ESTIMATED", "Use cautiously to avoid double counting."],
        ["Commercial Contingency", "% of direct cost", 0.07, "ESTIMATED", "Risk-based."],
        ["Overhead", "% of subtotal", 0.12, "ESTIMATED", "Company overhead."],
        ["Profit", "% of subtotal", 0.10, "ESTIMATED", "Management decision."],
    ]
    write_table(ws, 4, headers, rows, "tbl_indirects_risk")
    autofit(ws)

    cal = wb.create_sheet("EGYPT_CALIBRATION")
    hide_grid(cal)
    style_title(cal, "Egypt Calibration", "Editable project calibration factors")
    cal_headers = ["Factor", "Value", "Min", "Max", "Source_Status", "Notes"]
    cal_rows = [
        ["Labor_Factor", 1.00, 0.70, 1.50, "ESTIMATED", "Adjust for wage market and manpower availability."],
        ["Material_Factor", 1.00, 0.80, 1.70, "ESTIMATED", "Adjust for inflation/supplier quotes."],
        ["Equipment_Factor", 1.00, 0.80, 1.60, "ESTIMATED", "Adjust for equipment rental market."],
        ["Productivity_Factor", 1.00, 0.50, 1.30, "ESTIMATED", "Less than 1 reduces productivity."],
        ["Logistics_Factor", 1.00, 0.80, 1.80, "ESTIMATED", "Remote/offshore/desert access."],
        ["Risk_Factor", 1.00, 0.80, 1.50, "ESTIMATED", "Contract/site risk."],
    ]
    write_table(cal, 4, cal_headers, cal_rows, "tbl_egypt_calibration")
    autofit(cal)


def build_cost_engine(wb: Workbook) -> None:
    ws = wb.create_sheet("COST_ENGINE")
    hide_grid(ws)
    style_title(ws, "Cost Engine", "Auditable calculation view linked to BOQ and libraries")

    headers = [
        "Item_ID", "Discipline", "Activity_Code", "UOM", "Qty", "Productivity_Code", "Adj_Productivity_Day",
        "Crew_Code", "Crew_Cost_Day", "Crew_MH_Day", "Labor_Days", "Labor_Cost",
        "Equipment_Code", "Equipment_Daily_Cost", "Equipment_Days", "Equipment_Cost",
        "Material_Code", "Material_Unit_Cost", "Waste_%", "Material_Cost",
        "Direct_Cost", "Indirect_Adder", "Risk_Adder", "OH_Adder", "Profit_Adder", "Total_Cost", "Unit_Rate",
        "Flags"
    ]

    rows = []
    for idx in range(5, 205):
        boq_row = idx
        # formulas intentionally reference BOQ_INPUT and library tables using INDEX/MATCH for broad Excel compatibility
        rows.append([
            f"=IF(BOQ_INPUT!A{boq_row}=\"\",\"\",BOQ_INPUT!A{boq_row})",
            f"=IF(A{idx}=\"\",\"\",BOQ_INPUT!C{boq_row})",
            f"=IF(A{idx}=\"\",\"\",BOQ_INPUT!E{boq_row})",
            f"=IF(A{idx}=\"\",\"\",BOQ_INPUT!G{boq_row})",
            f"=IF(A{idx}=\"\",\"\",BOQ_INPUT!H{boq_row})",
            f"=IFERROR(INDEX(ACTIVITY_LIBRARY!G:G,MATCH(C{idx},ACTIVITY_LIBRARY!A:A,0)),\"[MISSING]\")",
            f"=IFERROR(INDEX(PRODUCTIVITY_LIBRARY!J:J,MATCH(F{idx},PRODUCTIVITY_LIBRARY!A:A,0))*INDEX(EGYPT_CALIBRATION!B:B,MATCH(\"Productivity_Factor\",EGYPT_CALIBRATION!A:A,0)),0)",
            f"=IFERROR(INDEX(ACTIVITY_LIBRARY!H:H,MATCH(C{idx},ACTIVITY_LIBRARY!A:A,0)),\"[MISSING]\")",
            f"=IFERROR(SUMIF(CREW_LIBRARY!A:A,H{idx},CREW_LIBRARY!G:G)*INDEX(EGYPT_CALIBRATION!B:B,MATCH(\"Labor_Factor\",EGYPT_CALIBRATION!A:A,0)),0)",
            f"=IFERROR(SUMIF(CREW_LIBRARY!A:A,H{idx},CREW_LIBRARY!H:H),0)",
            f"=IFERROR(E{idx}/G{idx},0)",
            f"=K{idx}*I{idx}",
            f"=IFERROR(INDEX(ACTIVITY_LIBRARY!I:I,MATCH(C{idx},ACTIVITY_LIBRARY!A:A,0)),\"[MISSING]\")",
            f"=IFERROR(INDEX(EQUIPMENT_LIBRARY!D:D,MATCH(M{idx},EQUIPMENT_LIBRARY!A:A,0))*INDEX(EGYPT_CALIBRATION!B:B,MATCH(\"Equipment_Factor\",EGYPT_CALIBRATION!A:A,0)),0)",
            f"=K{idx}",
            f"=N{idx}*O{idx}",
            f"=IFERROR(INDEX(ACTIVITY_LIBRARY!J:J,MATCH(C{idx},ACTIVITY_LIBRARY!A:A,0)),\"[MISSING]\")",
            f"=IFERROR(INDEX(MATERIAL_LIBRARY!D:D,MATCH(Q{idx},MATERIAL_LIBRARY!A:A,0))*INDEX(EGYPT_CALIBRATION!B:B,MATCH(\"Material_Factor\",EGYPT_CALIBRATION!A:A,0)),0)",
            f"=IFERROR(INDEX(MATERIAL_LIBRARY!E:E,MATCH(Q{idx},MATERIAL_LIBRARY!A:A,0)),0)",
            f"=E{idx}*R{idx}*(1+S{idx})",
            f"=L{idx}+P{idx}+T{idx}",
            f"=U{idx}*SUM(INDIRECTS_RISK!C5:C9)",
            f"=U{idx}*INDEX(EGYPT_CALIBRATION!B:B,MATCH(\"Risk_Factor\",EGYPT_CALIBRATION!A:A,0))*INDIRECTS_RISK!C10",
            f"=(U{idx}+V{idx}+W{idx})*INDIRECTS_RISK!C11",
            f"=(U{idx}+V{idx}+W{idx}+X{idx})*INDIRECTS_RISK!C12",
            f"=U{idx}+V{idx}+W{idx}+X{idx}+Y{idx}",
            f"=IFERROR(Z{idx}/E{idx},0)",
            f"=IF(A{idx}=\"\",\"\",IF(OR(C{idx}=\"[MISSING]\",F{idx}=\"[MISSING]\",H{idx}=\"[MISSING]\",Q{idx}=\"[MISSING]\"),\"REVIEW REQUIRED\",IF(Z{idx}=0,\"ZERO COST CHECK\",\"OK\")))",
        ])

    write_table(ws, 4, headers, rows, "tbl_cost_engine")
    freeze_and_filter(ws, 5)
    autofit(ws, max_width=28)

    # number formats
    for row in range(5, 205):
        for col in range(5, 28):
            ws.cell(row, col).number_format = '#,##0.00'


def build_outputs(wb: Workbook) -> None:
    ws = wb.create_sheet("OUTPUT_SUMMARY")
    hide_grid(ws)
    style_title(ws, "Output Summary", "Executive pricing summary")

    headers = ["Metric", "Value", "Notes"]
    rows = [
        ["Total BOQ Items", "=COUNTA(BOQ_INPUT!A5:A500)", "Number of input items."],
        ["Total Direct Cost", "=SUM(COST_ENGINE!U5:U204)", "Labor + equipment + material."],
        ["Total Indirects", "=SUM(COST_ENGINE!V5:V204)", "Project indirect adders."],
        ["Total Risk", "=SUM(COST_ENGINE!W5:W204)", "Risk adder."],
        ["Total OH", "=SUM(COST_ENGINE!X5:X204)", "Overhead."],
        ["Total Profit", "=SUM(COST_ENGINE!Y5:Y204)", "Profit."],
        ["Grand Total", "=SUM(COST_ENGINE!Z5:Z204)", "Commercial selling total before tax treatment."],
        ["Items Requiring Review", '=COUNTIF(COST_ENGINE!AB5:AB204,"REVIEW REQUIRED")', "Must close before submission."],
        ["Zero Cost Checks", '=COUNTIF(COST_ENGINE!AB5:AB204,"ZERO COST CHECK")', "Potential missing values."],
    ]
    write_table(ws, 4, headers, rows, "tbl_output_summary")
    for row in range(5, 14):
        ws.cell(row, 2).number_format = '#,##0.00'

    # Chart: cost breakdown
    chart = BarChart()
    chart.title = "Cost Breakdown"
    chart.y_axis.title = CURRENCY
    chart.x_axis.title = "Cost Element"
    data = Reference(ws, min_col=2, min_row=6, max_row=11)
    cats = Reference(ws, min_col=1, min_row=6, max_row=11)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, "E4")

    pie = PieChart()
    pie.title = "Pricing Composition"
    data2 = Reference(ws, min_col=2, min_row=6, max_row=11)
    cats2 = Reference(ws, min_col=1, min_row=6, max_row=11)
    pie.add_data(data2, titles_from_data=False)
    pie.set_categories(cats2)
    pie.height = 8
    pie.width = 10
    ws.add_chart(pie, "E20")

    autofit(ws)


def build_review_flags(wb: Workbook) -> None:
    ws = wb.create_sheet("REVIEW_FLAGS")
    hide_grid(ws)
    style_title(ws, "Review Flags", "Control sheet for estimator review before quote submission")

    headers = ["Check_ID", "Check", "Formula / Criteria", "Result", "Action Required"]
    rows = [
        ["CHK-001", "Missing project name", "PROJECT_MASTER Project Name = [MISSING]", '=IF(PROJECT_MASTER!B5="[MISSING]","FAIL","OK")', "Fill project name."],
        ["CHK-002", "BOQ has quantities", "BOQ qty count > 0", '=IF(COUNTA(BOQ_INPUT!H5:H500)>0,"OK","FAIL")', "Paste BOQ quantities."],
        ["CHK-003", "Mapped activity codes", "No blank activity codes", '=IF(COUNTBLANK(BOQ_INPUT!E5:E50)>40,"FAIL","OK")', "Map activities."],
        ["CHK-004", "No review required items", "Cost Engine flags", '=IF(COUNTIF(COST_ENGINE!AB5:AB204,"REVIEW REQUIRED")=0,"OK","FAIL")', "Close missing mappings/rates."],
        ["CHK-005", "No zero cost items", "Zero cost flags", '=IF(COUNTIF(COST_ENGINE!AB5:AB204,"ZERO COST CHECK")=0,"OK","FAIL")', "Check missing rates/productivity."],
        ["CHK-006", "Material missing count", "MATERIAL_LIBRARY missing", '=COUNTIF(MATERIAL_LIBRARY!F5:F100,"MISSING")', "Replace missing material costs."],
        ["CHK-007", "Equipment missing count", "EQUIPMENT_LIBRARY missing", '=COUNTIF(EQUIPMENT_LIBRARY!F5:F100,"MISSING")', "Replace missing equipment costs."],
        ["CHK-008", "Productivity missing count", "PRODUCTIVITY_LIBRARY missing", '=COUNTIF(PRODUCTIVITY_LIBRARY!K5:K100,"MISSING")', "Replace missing productivity."],
        ["CHK-009", "Commercial summary total", "Grand total > 0", '=IF(OUTPUT_SUMMARY!B11>0,"OK","FAIL")', "Engine must calculate total."],
    ]
    write_table(ws, 4, headers, rows, "tbl_review_flags")
    autofit(ws)


def build_source_register_and_log(wb: Workbook) -> None:
    src = wb.create_sheet("SOURCE_REGISTER")
    hide_grid(src)
    style_title(src, "Source Register", "Evidence trail for all assumptions and rates")
    headers = ["Source_ID", "Category", "Description", "Status", "Owner", "Date", "Reference_Link_or_Doc", "Used_In", "Notes"]
    rows = [
        ["SRC-001", "Labor Rates", "Local labor daily rates", "ESTIMATED", "Estimator", "=TODAY()", "Internal assumption", "CREW_LIBRARY", "Replace with payroll/vendor data."],
        ["SRC-002", "Equipment", "Equipment rental daily rates", "ESTIMATED", "Estimator", "=TODAY()", "Supplier quote required", "EQUIPMENT_LIBRARY", "Get signed quotation."],
        ["SRC-003", "Materials", "Consumables and materials", "MISSING/ESTIMATED", "Procurement", "=TODAY()", "Supplier quote required", "MATERIAL_LIBRARY", "Attach quotations before bid."],
        ["SRC-004", "Productivity", "Crew productivity norms", "ESTIMATED", "Construction", "=TODAY()", "Company norm required", "PRODUCTIVITY_LIBRARY", "Calibrate with actual jobs."],
        ["SRC-005", "Indirects", "Project indirect percentages", "ESTIMATED", "Management", "=TODAY()", "Management decision", "INDIRECTS_RISK", "Review by project type."],
    ]
    write_table(src, 4, headers, rows, "tbl_source_register")
    autofit(src)

    log = wb.create_sheet("CHANGE_LOG")
    hide_grid(log)
    style_title(log, "Change Log", "Version control and audit trail")
    log_headers = ["Revision", "Date", "Changed_By", "Change_Description", "Reason", "Approved_By"]
    log_rows = [
        ["V1.0", "=TODAY()", "ChatGPT / PETROCAF", "Initial pricing workbook generator created", "Build GitHub starter engine", "Ahmed Tarakhan"],
    ]
    write_table(log, 4, log_headers, log_rows, "tbl_change_log")
    autofit(log)


def final_format(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    # Make cost engine and library sheets visible but ordered logically
    order = [
        "README_BASIS", "PROJECT_MASTER", "BOQ_INPUT", "OUTPUT_SUMMARY", "REVIEW_FLAGS",
        "COST_ENGINE", "ACTIVITY_LIBRARY", "PRODUCTIVITY_LIBRARY", "CREW_LIBRARY",
        "MATERIAL_LIBRARY", "EQUIPMENT_LIBRARY", "INDIRECTS_RISK", "EGYPT_CALIBRATION",
        "SOURCE_REGISTER", "CHANGE_LOG",
    ]
    wb._sheets = [wb[s] for s in order if s in wb.sheetnames]


def build_workbook(output_path: str | Path = OUTPUT_FILE) -> Path:
    wb = Workbook()
    build_readme(wb)
    build_project_master(wb)
    build_boq_input(wb)
    build_activity_library(wb)
    build_productivity_library(wb)
    build_crew_library(wb)
    build_cost_libraries(wb)
    build_indirects_and_calibration(wb)
    build_cost_engine(wb)
    build_outputs(wb)
    build_review_flags(wb)
    build_source_register_and_log(wb)
    final_format(wb)

    path = Path(output_path)
    wb.save(path)
    return path


if __name__ == "__main__":
    output = build_workbook()
    print(f"Created: {output.resolve()}")
